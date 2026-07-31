"""Excel builders for settlement / revenue-sharing reports."""
from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font


def settlements_workbook(settlements: Iterable) -> BytesIO:
    wb = Workbook()
    summary = wb.active
    summary.title = 'Settlements'
    summary.append([
        'Settlement ID', 'Technician', 'Mobile', 'Period Start', 'Period End',
        'Cadence', 'Status', 'Gross', 'Incentive', 'Deduction', 'Net', 'Lines',
    ])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    lines_sheet = wb.create_sheet('Line Items')
    lines_sheet.append([
        'Settlement ID', 'Technician', 'Job Code', 'Earning Type', 'Amount', 'Notes',
    ])
    for cell in lines_sheet[1]:
        cell.font = Font(bold=True)

    for s in settlements:
        summary.append([
            s.id,
            s.technician.name,
            s.technician.mobile,
            s.period_start.isoformat(),
            s.period_end.isoformat(),
            s.cadence,
            s.status,
            float(s.gross_amount),
            float(s.incentive_amount),
            float(s.deduction_amount),
            float(s.net_amount),
            s.line_items.count(),
        ])
        for line in s.line_items.select_related('job').all():
            lines_sheet.append([
                s.id,
                s.technician.name,
                line.job.code,
                line.earning_type,
                float(line.amount),
                line.notes or '',
            ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def revenue_sharing_workbook(rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Revenue Sharing'
    headers = [
        'Job Code', 'Completed At', 'Technician', 'Partner', 'Payment Model',
        'Payout Status', 'Visit Revenue', 'Tech Pool', 'Company Share',
        'Visit Payout', 'Earning Amount', 'Earning Approved',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h) for h in [
            'job_code', 'completed_at', 'technician', 'partner', 'payment_model',
            'payout_status', 'visit_revenue', 'tech_pool', 'company_share',
            'visit_payout', 'earning_amount', 'earning_approved',
        ]])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
